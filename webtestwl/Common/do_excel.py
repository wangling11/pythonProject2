# -- coding: utf-8 --
# @Time : 2022/5/26 16:50
# @Author : wangling
# @Email : 731569578@qq.com
# @File : do_excel.py
# @Software: PyCharm
"""
excel类，你的需求是实现是什么?
1、读取表头
2、读取数据 - 读取表头以外的所有数据。 - 返回值：列表，成员是每一行数据

初始化工作？  加载一个excel,打开一个表单。

"""
import os

from openpyxl import load_workbook;
# 读取Excel里面的数据
class Do_Excel:

    def __init__(self,filename,sheetname):
        self.wb = load_workbook(filename)
        self.sh = self.wb[sheetname]

    """读取表单的第一行的每个列"""
    def __read_titles(self):
        titles = []
        for item in list(self.sh.rows)[0]:  # 遍历第1行当中每一列
            titles.append(item.value)
        return titles;

    def read_all_datas(self):
        all_datas = []
        titles = self.__read_titles()
        #得到最大化
        for item in list(self.sh.rows)[1:]:  # 遍历数据行
            values = []
            for val in item:  # 获取每一行的值
                values.append(val.value)
            res = dict(zip(titles, values))  # title和每一行数据，打包成字典
            all_datas.append(res)
        return all_datas


    def close_file(self):
        self.wb.close()

    def getData(self,filename,sheetname ):

        # 1 打开excel
        wb = load_workbook(filename)
        # 2 定位表单
        sheet = wb[sheetname];
        # 3 定位单元格  行列值
        res = sheet.cell(1, 1).value;
        print(sheet.max_row);  # 最大行
        print(sheet.max_column);  # 最大列
        print(res);  # 最大列

if __name__ == '__main__':
    from do_path import datas_dir;
    filename = os.path.join(datas_dir,"test.xlsx")
    #os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    print(filename)
    exc=Do_Excel(filename,"register");
    cases = exc.read_all_datas()
    exc.close_file()
    for case in cases:
        print(case)
